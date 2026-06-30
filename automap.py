#!/usr/bin/env python3
"""Automap — M&A Company Mapping Tool · Accomplir Advisors"""

import os
import json
import time
import re
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional

import requests
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

# ── Config ────────────────────────────────────────────────────────────────────
SERP_API_KEY   = os.getenv("SERP_API_KEY", "")
APOLLO_API_KEY = os.getenv("APOLLO_API_KEY", "")
CLAUDE_MODEL   = "claude-sonnet-4-6"
MAX_COMPANIES_PER_CATEGORY = 60
SERP_QUERIES_PER_CATEGORY  = 5
SERP_AVAILABLE = True             # flipped to False on first 429

claude = anthropic.Anthropic()    # reads ANTHROPIC_API_KEY from env

# ── Excel theme ───────────────────────────────────────────────────────────────
NAVY       = "1F3864"
BLUE       = "2E75B6"
LIGHT_BLUE = "EEF2F7"
WHITE      = "FFFFFF"
GRAY_TEXT  = "808080"
DARK_TEXT  = "404040"
BORDER_COL = "BDD7EE"

def _side():
    return Side(style="thin", color=BORDER_COL)

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style(cell, size=9):
    cell.font      = Font(name="Arial", size=size, bold=True, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()

def _data_style(cell, bold=False, align="left", row_idx=0):
    bg = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    cell.border    = _border()

# ── SerpAPI ───────────────────────────────────────────────────────────────────
def _serp(query: str, num: int = 10) -> List[Dict]:
    global SERP_AVAILABLE
    if not SERP_AVAILABLE or not SERP_API_KEY:
        return []
    try:
        r = requests.get(
            "https://serpapi.com/search",
            params={"q": query, "api_key": SERP_API_KEY, "num": num, "gl": "in", "hl": "en"},
            timeout=15,
        )
        if r.status_code == 429:
            SERP_AVAILABLE = False
            print("    [SerpAPI credits exhausted — enriching from Claude knowledge only]")
            return []
        r.raise_for_status()
        return r.json().get("organic_results", [])
    except Exception as e:
        print(f"    [search error] {e}")
        return []

# ── Apollo — Step 2: decision maker names (free, no credits) ─────────────────

# Titles we want: promoters, C-suite, M&A/Corp Dev heads
_DM_KEEP = [
    "ceo", "chief executive",
    "cfo", "chief financial",
    "coo", "chief operating",
    "cmd", "chairman", "vice chairman",
    "founder", "co-founder", "promoter", "owner",
    "managing director", "joint managing director", "jmd",
    "executive director",
    "m&a", "merger", "acquisition", "corporate development", "corp dev",
    "group president", "president",
    "managing partner", "principal",
]

# If any of these appear in the title, discard even if above matched
_DM_EXCLUDE = [
    "hr ", "h.r.", "human resource",
    "purchase", "procurement", "supply chain",
    "sales", "marketing", "business development",
    "information technology", " it ", "legal", "compliance",
    "admin", "avp", "assistant vice",
    "technical", "engineering", "production", "quality",
]

def _relevant_dm(title: str) -> bool:
    """Return True only for owners, promoters, C-suite, and M&A-specific roles."""
    t = title.lower()
    if not any(k in t for k in _DM_KEEP):
        return False
    if any(k in t for k in _DM_EXCLUDE):
        return False
    return True

def apollo_people(company_name: str) -> List[Dict]:
    """Return owners / C-suite / M&A heads at a company. Zero credits consumed."""
    if not APOLLO_API_KEY:
        return []
    try:
        params = [
            ("q_keywords",            company_name),
            ("person_seniorities[]",  "c_suite"),
            ("person_seniorities[]",  "founder"),
            ("person_seniorities[]",  "owner"),
            ("person_seniorities[]",  "director"),   # catches MD, ED, JMD in Indian companies
            ("per_page",              15),            # fetch more, then filter down
            ("page",                  1),
        ]
        r = requests.post(
            "https://api.apollo.io/api/v1/mixed_people/api_search",
            headers={"x-api-key": APOLLO_API_KEY, "Content-Type": "application/json"},
            params=params,
            timeout=15,
        )
        r.raise_for_status()

        people = r.json().get("people", [])
        relevant = [p for p in people if _relevant_dm(p.get("title", ""))]

        # If strict filter returns nothing, fall back to c_suite / founder / owner only
        if not relevant:
            relevant = [
                p for p in people
                if p.get("seniority") in ("c_suite", "founder", "owner")
            ]

        return [
            {
                "apollo_id":            p.get("id", ""),
                "first_name":           p.get("first_name", ""),
                "last_name_obfuscated": p.get("last_name_obfuscated", ""),
                "title":                p.get("title", ""),
                "has_email":            "Y" if p.get("has_email") else "N",
                "has_phone":            "Y" if p.get("has_direct_phone") == "Yes" else "N",
            }
            for p in relevant[:5]   # cap at 5 per company
        ]
    except Exception as e:
        print(f"    [Apollo error] {e}")
        return []

# ── Claude helpers ────────────────────────────────────────────────────────────
def _ask(prompt: str, system: str = "", max_tokens: int = 4096) -> str:
    resp = claude.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=max_tokens,
        system=system or "You are a senior M&A analyst at Accomplir Advisors, an Indian investment bank.",
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.content[0].text

def _json(text: str) -> Any:
    """Extract JSON from a Claude response that may contain markdown fences."""
    text = text.strip()
    if "```json" in text:
        text = text.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in text:
        text = text.split("```", 1)[1].split("```", 1)[0].strip()
    # Find outermost JSON object or array
    m = re.search(r"(\{.*\}|\[.*\])", text, re.DOTALL)
    if m:
        text = m.group(1)
    return json.loads(text)

# ── Input collection ──────────────────────────────────────────────────────────
MOTIVATIONS = [
    "Capacity expansion",
    "Geographic diversification",
    "Product / service diversification",
    "Backward integration",
    "Forward integration",
    "Technology / IP acquisition",
    "Financial / PE buy-out",
    "Acqui-hire (talent acquisition)",
]

def _require(prompt: str, example: str = "") -> str:
    """Prompt until non-empty input is given."""
    hint = f" (e.g. '{example}')" if example else ""
    while True:
        val = input(f"{prompt}{hint}: ").strip()
        if val:
            return val
        print("  ↳ Required — please enter a value.")

def _multiline(label: str) -> str:
    """Collect multi-line input; blank line terminates."""
    print(f"{label}")
    print("  (Paste as many lines as you like. Press Enter on a blank line when done.)")
    lines = []
    while True:
        line = input()
        if line.strip() == "":
            if lines:
                break
        else:
            lines.append(line)
    return " ".join(lines)

COMPANY_TYPES = [
    "Listed (BSE / NSE)",
    "Unlisted / private",
    "PE / VC backed",
    "Family-owned / promoter-driven",
    "MNC subsidiary",
    "PSU / Government",
]

COUNT_OPTIONS = {
    "1": ("~100",         15),   # per-category cap → ~100-120 total
    "2": ("~300",         40),
    "3": ("~500",         65),
    "4": ("~1000",       130),
    "5": ("Full universe — no cap", 250),
}

def _optional(prompt: str, example: str = "") -> str:
    """Optional input — blank is fine."""
    hint = f" (e.g. '{example}')" if example else ""
    return input(f"{prompt}{hint} [Enter to skip]: ").strip()

def _multiselect(options: List[str], label: str) -> List[str]:
    """Multi-select from a numbered list. Blank = all."""
    print(f"\n{label} — enter numbers, or press Enter for all:")
    for i, o in enumerate(options, 1):
        print(f"  {i}. {o}")
    raw = input("Selection [or Enter for all]: ").strip()
    if not raw:
        return options[:]
    idxs = []
    for tok in re.split(r"[,\s]+", raw):
        if tok.isdigit():
            idxs.append(int(tok) - 1)
    chosen = [options[i] for i in idxs if 0 <= i < len(options)]
    return chosen if chosen else options[:]

def collect_inputs() -> Dict:
    print("\n" + "═" * 62)
    print("  AUTOMAP  ·  Accomplir Advisors  ·  M&A Mapping Tool  v1")
    print("═" * 62 + "\n")

    # ── Mandate type ──────────────────────────────────────────────────────────
    while True:
        choice = input("Mandate type:\n  1. Sell-side  (company looking to sell)\n  2. Buy-side   (company looking to acquire)\nChoose [1/2]: ").strip()
        if choice in ("1", "2"):
            break
    is_sellside = (choice == "1")

    client       = _require("\nClient / company name")
    role         = "the company being sold" if is_sellside else "the acquirer"
    company_desc = _multiline(f"\nDescribe {client} ({role}) — products, size, geography, any special context:")
    sector       = _require("\nSector", "Specialty Chemicals")
    sub_sector   = _require("Sub-sector / product focus", "Plasticizers & PVC compounds")
    geography    = _require("Client geography", "Pan-India / Gujarat")

    # ── M&A motivations ───────────────────────────────────────────────────────
    print("\nM&A motivations — enter numbers separated by commas:")
    for i, m in enumerate(MOTIVATIONS, 1):
        print(f"  {i}. {m}")
    while True:
        raw = input("Selection: ").strip()
        idxs = []
        for tok in re.split(r"[,\s]+", raw):
            if tok.isdigit():
                idxs.append(int(tok) - 1)
        motivations = [MOTIVATIONS[i] for i in idxs if 0 <= i < len(MOTIVATIONS)]
        if motivations:
            break
        print("  ↳ Please enter at least one valid number from the list.")

    # ── Target company filters ────────────────────────────────────────────────
    print("\n" + "─" * 62)
    print("  TARGET COMPANY FILTERS  (all optional — press Enter to skip)")
    print("─" * 62)

    target_geography = _optional(
        "Geography of companies to map",
        "Pan-India" if is_sellside else "South India, Maharashtra"
    )
    revenue_range = _optional(
        "Revenue range of target companies",
        ">100 Cr  or  50–500 Cr  or  <50 Cr"
    )
    specific_attrs = _optional(
        "Specific attributes target companies must / should have",
        "manufacturing plant in India  /  CRISIL rated  /  has export revenue"
    )
    what_looking_for = _optional(
        "What is the buyer / seller specifically looking for in the target",
        "geographic presence in East India  /  R&D capabilities  /  PLI beneficiary"
    )

    company_types = _multiselect(COMPANY_TYPES, "Company types to include")

    # ── How many companies ────────────────────────────────────────────────────
    print("\nHow many companies to map?")
    for k, (label, _) in COUNT_OPTIONS.items():
        print(f"  {k}. {label}")
    while True:
        cnt = input("Choose [1-5]: ").strip()
        if cnt in COUNT_OPTIONS:
            break
        print("  ↳ Enter 1–5.")
    count_label, per_cat_cap = COUNT_OPTIONS[cnt]

    return {
        "is_sellside":       is_sellside,
        "client":            client,
        "company_desc":      company_desc,
        "sector":            sector,
        "sub_sector":        sub_sector,
        "geography":         geography,
        "motivations":       motivations,
        # filters
        "target_geography":  target_geography  or "Pan-India",
        "revenue_range":     revenue_range     or "No filter",
        "specific_attrs":    specific_attrs    or "None",
        "what_looking_for":  what_looking_for  or "None",
        "company_types":     company_types,
        "per_cat_cap":       per_cat_cap,
        "count_label":       count_label,
        "date":              datetime.today().strftime("%d-%b-%y"),
    }

# ── Strategy generation — two-step to avoid token truncation ─────────────────
def generate_strategy(inp: Dict) -> Dict:
    """Step 1: Get categories + rationale + search queries (no companies yet)."""
    side   = "sell-side (company looking to sell)" if inp["is_sellside"] else "buy-side (company looking to acquire)"
    target = "potential buyers / acquirers" if inp["is_sellside"] else "potential acquisition targets"

    filters = f"""
Target company filters:
  Geography:          {inp['target_geography']}
  Revenue range:      {inp['revenue_range']}
  Company types:      {', '.join(inp['company_types'])}
  Must-have attributes: {inp['specific_attrs']}
  Buyer/seller seeking: {inp['what_looking_for']}"""

    prompt = f"""Mandate: {side}
Client: {inp['client']}
Company description: {inp['company_desc']}
Sector: {inp['sector']}  |  Sub-sector: {inp['sub_sector']}
Client geography: {inp['geography']}
M&A motivations: {', '.join(inp['motivations'])}
{filters}

Identify 6–8 distinct, non-overlapping categories of {target} in India that match the filters above.
For each category also provide 6 targeted Google search queries to find regional / unlisted / smaller players — vary by state, product sub-type, company size, and industry directories.
Tailor queries to the target geography specified above.

Return ONLY valid JSON:
{{
  "mandate_summary": "2-3 sentence summary",
  "categories": [
    {{
      "name": "Category name (under 8 words)",
      "description": "What types of companies fall here and why relevant",
      "rationale": "Why these companies would pursue this M&A — specific to {inp['client']} and the stated motivations (2-3 sentences)",
      "search_queries": ["query 1", "query 2", "query 3", "query 4", "query 5", "query 6"]
    }}
  ]
}}"""

    print("\n[1/4] Generating mapping strategy...")
    return _json(_ask(prompt, max_tokens=4000))


def _companies_for_category(cat: Dict, inp: Dict) -> List[Dict]:
    """Ask Claude for companies in one category — respects filters and count target."""
    side = "sell-side" if inp["is_sellside"] else "buy-side"
    cap  = inp.get("per_cat_cap", 65)

    # Build filter block — only show non-trivial filters to keep prompt focused
    filter_lines = []
    if inp["target_geography"] != "Pan-India":
        filter_lines.append(f"  Geography: {inp['target_geography']} (prioritise, but include Pan-India players too if relevant)")
    if inp["revenue_range"] != "No filter":
        filter_lines.append(f"  Revenue:   {inp['revenue_range']}")
    types = inp.get("company_types", [])
    if types and len(types) < len(COMPANY_TYPES):
        filter_lines.append(f"  Types:     {', '.join(types)}")
    if inp["specific_attrs"] != "None":
        filter_lines.append(f"  Must have: {inp['specific_attrs']}")
    if inp["what_looking_for"] != "None":
        filter_lines.append(f"  Buyer/seller seeking: {inp['what_looking_for']}")
    filter_block = ("\nFilters to respect:\n" + "\n".join(filter_lines)) if filter_lines else ""

    universe_note = (
        "This is a full-universe sweep — list every company you know, no matter how small or regional."
        if cap >= 200
        else f"Target: list up to {cap} companies. Cast a wide net — include smaller and regional players."
    )

    prompt = f"""Mandate: {side} for {inp['client']} ({inp['sub_sector']})

Category: {cat['name']}
Description: {cat['description']}
{filter_block}

{universe_note}
Include:
  • Large listed AND mid-size private/unlisted companies
  • Pan-India AND state/regional players
  • Family-owned, PE-backed, JVs, subsidiaries, PSUs/cooperatives
Do NOT invent names. Only include companies you are confident exist.
Exclude {inp['client']} itself.

Return ONLY valid JSON:
{{"companies": [{{"name": "Full company name", "location": "City, State", "snippet": "one sentence: core business and scale"}}]}}"""

    try:
        result = _json(_ask(prompt, max_tokens=8000))
        return result.get("companies", [])
    except Exception as e:
        print(f"    [company list error for {cat['name']}] {e}")
        return []


# ── Company discovery — Claude list + SerpAPI sweep ───────────────────────────
def discover_companies(categories: List[Dict], inp: Dict) -> Dict[str, List[Dict]]:
    results = {}

    for i, cat in enumerate(categories, 1):
        cat_name = cat["name"]
        print(f"\n  [{i}/{len(categories)}] {cat_name}")

        # Claude generates companies for this category
        claude_cos = _companies_for_category(cat, inp)
        seen: dict = {}
        for co in claude_cos:
            key = co["name"].lower().strip()
            seen[key] = co

        print(f"    Claude: {len(seen)} companies", end="", flush=True)

        # SerpAPI sweep — finds regional/unlisted players Claude may not know
        if SERP_AVAILABLE:
            for q in cat.get("search_queries", [])[:SERP_QUERIES_PER_CATEGORY]:
                hits = _serp(q, num=10)
                if not hits:
                    continue

                snippets = "\n".join(
                    f"- {h.get('title','')} | {h.get('link','')} | {h.get('snippet','')}"
                    for h in hits
                )
                parse_prompt = f"""From these search results, extract Indian company names in the {inp['sector']} sector.

Query: {q}
Results:
{snippets}

Return ONLY valid JSON:
{{"companies": [{{"name": "Full company name", "location": "City, State or null", "snippet": "one sentence on what they do"}}]}}

Rules:
- Only real Indian companies (incorporated in India or significant India operations)
- Exclude directories, news articles, associations, analyst reports
- Maximum 8 companies per result set"""

                try:
                    parsed = _json(_ask(parse_prompt))
                    for co in parsed.get("companies", []):
                        key = co["name"].lower().strip()
                        if key not in seen:
                            seen[key] = co
                except Exception:
                    pass
                time.sleep(0.3)

            added = len(seen) - len(claude_cos)
            print(f" + {added} from search = {len(seen)} total")
        else:
            print()

        all_cos = list(seen.values())[:inp.get("per_cat_cap", MAX_COMPANIES_PER_CATEGORY)]
        results[cat_name] = all_cos

    return results

# ── Company enrichment ────────────────────────────────────────────────────────
def _enrich_one(company: Dict, category: str, inp: Dict) -> Dict:
    name = company["name"]

    # Single combined search — saves 2/3 of SerpAPI credits vs separate searches
    combined_hits = _serp(
        f"{name} India revenue EBITDA credit rating CEO MD annual report FY25", num=10
    )
    snippets = "\n".join(
        f"- {h.get('title','')} | {h.get('snippet','')}"
        for h in combined_hits
    )
    if not snippets:
        snippets = "(No live search results — use Claude knowledge only for description and rationale; set all financial figures to null)"

    rationale_q = (
        f"Why would {name} want to acquire {inp['client']}?"
        if inp["is_sellside"]
        else f"Why would {inp['client']} want to acquire {name}?"
    )

    prompt = f"""Extract information about the Indian company "{name}" from these search results.

Category: {category}
Location hint: {company.get('location', 'India')}
Mandate context ({('sell-side' if inp['is_sellside'] else 'buy-side')}): {inp['company_desc']}
Motivations: {', '.join(inp['motivations'])}

Search results:
{snippets}

Return ONLY valid JSON (use null for anything not found — never guess financials):
{{
  "description": "2–3 sentence company overview: founding, core business, scale, key clients/sectors",
  "products_services": "key products or services, one per line",
  "listed": "Y or N or Unknown",
  "revenue": null or number in INR Crore (latest year),
  "ebitda": null or number in INR Crore,
  "ebitda_pct": null or decimal (0.15 = 15%),
  "pat": null or number in INR Crore,
  "pat_pct": null or decimal,
  "market_cap": null or number in INR Crore (listed companies only),
  "debt": null or number in INR Crore,
  "cash": null or number in INR Crore,
  "financial_year": null or "FY25" / "FY24",
  "peak_revenue": null or "X.X (FYxx)" string,
  "credit_rating": null or "AGENCY RATING; Outlook" or "Not rated",
  "management": "Name - Title (Age if known)\\nName - Title",
  "shareholding": "promoter %  |  public % or key institutional holders",
  "strategic_rationale": "2–3 sentences answering: {rationale_q}"
}}

CRITICAL: Only include financial figures explicitly stated in the results. Do not estimate or hallucinate numbers."""

    try:
        result = _json(_ask(prompt))
    except Exception as e:
        print(f"    [enrich error: {name}] {e}")
        result = {}

    # Compute EV and multiples
    mc    = result.get("market_cap")
    debt  = result.get("debt")
    cash  = result.get("cash")
    rev   = result.get("revenue")
    ebitda = result.get("ebitda")

    ev = None
    if mc is not None and debt is not None:
        ev = mc + debt - (cash or 0)
    result["ev"] = ev

    result["ttm_ev_rev"]    = round(ev / rev, 2)    if ev and rev             else None
    result["ttm_ev_ebitda"] = round(ev / ebitda, 2) if ev and ebitda and ebitda != 0 else None

    result["name"]     = name
    result["category"] = category
    result["location"] = company.get("location") or "India"

    # Apollo Step 2 — decision maker names (free, no credits)
    result["apollo_people"] = apollo_people(name)
    time.sleep(0.2)   # gentle rate limiting for Apollo

    return result

def enrich_all(discovered: Dict[str, List[Dict]], inp: Dict) -> Dict[str, List[Dict]]:
    total = sum(len(v) for v in discovered.values())
    done  = 0
    print(f"\n[3/4] Enriching {total} companies (descriptions + rationale via Claude; financials via search where available)...")

    enriched: Dict[str, List[Dict]] = {}
    for category, companies in discovered.items():
        enriched[category] = []
        for co in companies:
            done += 1
            print(f"  [{done}/{total}] {co['name']}")
            enriched[category].append(_enrich_one(co, category, inp))
            time.sleep(0.2)

    return enriched

# ── Excel: shared helpers ─────────────────────────────────────────────────────
def _write_header_block(ws, title: str, client: str, date_str: str):
    ws.row_dimensions[1].height = 5
    ws.cell(2, 2, "Accomplir Advisors").font = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws.cell(3, 2, f"Market mapping — {title}").font = Font(name="Arial", size=11, italic=True, color=BLUE)
    ws.cell(4, 2, f"For {client}").font = Font(name="Arial", size=10, color=DARK_TEXT)
    ws.cell(5, 2, date_str).font = Font(name="Arial", size=10, color=DARK_TEXT)
    ws.cell(6, 2, "INR in Cr unless specified").font = Font(name="Arial", size=9, italic=True, color=GRAY_TEXT)
    ws.row_dimensions[7].height = 8
    ws.row_dimensions[8].height = 5

# ── Excel: Mapping Snapshot sheet ────────────────────────────────────────────
SNAPSHOT_COLS = [
    ("S.\nNo.", 5),
    ("Category", 30),
    ("Description", 48),
    ("Rationale for Mapping", 58),
]

def _snapshot_sheet(wb: Workbook, strategy: Dict, inp: Dict):
    ws = wb.create_sheet("Mapping Snapshot")
    side_label = "sell-side" if inp["is_sellside"] else "buy-side"
    _write_header_block(ws, f"{inp['sub_sector']} — {side_label} mapping", inp["client"], inp["date"])

    ws.row_dimensions[9].height = 36
    for ci, (label, width) in enumerate(SNAPSHOT_COLS, 1):
        _header_style(ws.cell(9, ci, label))
        ws.column_dimensions[get_column_letter(ci)].width = width

    for i, cat in enumerate(strategy["categories"], 1):
        r = 9 + i
        ws.row_dimensions[r].height = 72
        for ci, val in enumerate([i, cat["name"], cat["description"], cat["rationale"]], 1):
            cell = ws.cell(r, ci, val)
            _data_style(cell, bold=(ci == 2), row_idx=i)

    ws.freeze_panes = "B10"

# ── Excel: Main mapping sheet ─────────────────────────────────────────────────
# (label, width, align, data_key)
MAIN_COLS = [
    ("S.\nNo.",               6,  "center", "_sno"),
    ("Company Name",         26,  "left",   "name"),
    ("Company\nCategory",    20,  "left",   "category"),
    ("Location (HO)",        18,  "left",   "location"),
    ("About the Company",    42,  "left",   "description"),
    ("Products /\nServices", 30,  "left",   "products_services"),
    ("Listed\n(Y/N)",        10,  "center", "listed"),
    ("Revenue\n(INR Cr)",    12,  "right",  "revenue"),
    ("EBITDA\n(INR Cr)",     12,  "right",  "ebitda"),
    ("EBITDA %",             10,  "right",  "ebitda_pct"),
    ("PAT\n(INR Cr)",        11,  "right",  "pat"),
    ("PAT %",                 9,  "right",  "pat_pct"),
    ("Market Cap\n(INR Cr)", 13,  "right",  "market_cap"),
    ("Debt\n(INR Cr)",       10,  "right",  "debt"),
    ("Cash\n(INR Cr)",       10,  "right",  "cash"),
    ("EV\n(INR Cr)",         11,  "right",  "ev"),
    ("Financials\nYear",     12,  "center", "financial_year"),
    ("Peak Revenue",         14,  "center", "peak_revenue"),
    ("Credit Rating",        20,  "left",   "credit_rating"),
    ("TTM EV /\nRev (x)",   11,  "right",  "ttm_ev_rev"),
    ("TTM EV /\nEBITDA (x)",13,  "right",  "ttm_ev_ebitda"),
    ("Management Profile",   36,  "left",   "management"),
    ("Shareholding\nSummary",28,  "left",   "shareholding"),
    ("Strategic Rationale /\nIntent", 42, "left", "strategic_rationale"),
    ("Notes",                22,  "left",   "_notes"),
    ("Client\nFeedback",     22,  "left",   "_feedback"),
]

PCT_FIELDS      = {"ebitda_pct", "pat_pct"}
MULTIPLE_FIELDS = {"ttm_ev_rev", "ttm_ev_ebitda"}

def _fmt(val, key: str):
    if val is None:
        return ""
    if key in PCT_FIELDS:
        return f"{val*100:.1f}%"
    if key in MULTIPLE_FIELDS:
        return f"{val:.1f}x"
    if isinstance(val, float):
        return round(val, 2)
    return val

def _mapping_sheet(wb: Workbook, name: str, companies: List[Dict],
                   title: str, client: str, date_str: str):
    safe = re.sub(r"[\/\\\?\*\[\]:]", "", name)[:31]
    ws   = wb.create_sheet(safe)

    _write_header_block(ws, title, client, date_str)

    # Column headers — row 9
    ws.row_dimensions[9].height = 40
    for ci, (label, width, _, __) in enumerate(MAIN_COLS, 1):
        _header_style(ws.cell(9, ci, label))
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for i, co in enumerate(companies, 1):
        r = 9 + i
        ws.row_dimensions[r].height = 90
        for ci, (_, __, align, key) in enumerate(MAIN_COLS, 1):
            if key == "_sno":
                val = i
            elif key in ("_notes", "_feedback"):
                val = ""
            else:
                val = _fmt(co.get(key), key)
            cell = ws.cell(r, ci, val)
            _data_style(cell, bold=(ci == 2), align=align, row_idx=i)

    ws.freeze_panes = "C10"
    last_col = get_column_letter(len(MAIN_COLS))
    ws.auto_filter.ref = f"A9:{last_col}9"

# ── Excel: Decision Makers sheet (Apollo Step 2 output) ──────────────────────
DM_COLS = [
    ("S.\nNo.",          5,  "center"),
    ("Company Name",    28,  "left"),
    ("Category",        20,  "left"),
    ("First Name",      14,  "left"),
    ("Last Name\n(partial)", 16, "left"),
    ("Full Name\n(complete manually)", 22, "left"),
    ("Title",           28,  "left"),
    ("Has\nEmail",       9,  "center"),
    ("Has\nPhone",       9,  "center"),
    ("Email",           30,  "left"),
    ("Phone",           18,  "left"),
    ("Enrich?\n(Y to pull email/phone)", 14, "center"),
    ("Apollo ID",       28,  "left"),
]

def _decision_makers_sheet(wb: Workbook, enriched: Dict[str, List[Dict]], date_str: str):
    ws = wb.create_sheet("Decision Makers")

    # Header block
    ws.row_dimensions[1].height = 5
    ws.cell(2, 2, "Accomplir Advisors").font = Font(name="Arial", size=14, bold=True, color=NAVY)
    ws.cell(3, 2, "Decision Makers — sourced via Apollo (Step 2)").font = Font(name="Arial", size=11, italic=True, color=BLUE)
    ws.cell(4, 2, "Last name is partially masked by Apollo. Complete 'Full Name' column manually, then mark 'Y' in Enrich column and run contacts.py to pull emails.").font = Font(name="Arial", size=9, italic=True, color=GRAY_TEXT)
    ws.cell(5, 2, date_str).font = Font(name="Arial", size=10, color=DARK_TEXT)
    ws.row_dimensions[7].height = 5

    # Column headers — row 8
    ws.row_dimensions[8].height = 40
    for ci, (label, width, _) in enumerate(DM_COLS, 1):
        _header_style(ws.cell(8, ci, label))
        ws.column_dimensions[get_column_letter(ci)].width = width

    row = 8
    sno = 0
    for category, companies in enriched.items():
        for co in companies:
            for person in co.get("apollo_people", []):
                sno += 1
                row += 1
                ws.row_dimensions[row].height = 20
                vals = [
                    sno,
                    co["name"],
                    category,
                    person.get("first_name", ""),
                    person.get("last_name_obfuscated", ""),
                    "",   # Full Name — manual
                    person.get("title", ""),
                    person.get("has_email", ""),
                    person.get("has_phone", ""),
                    "",   # Email — filled by contacts.py
                    "",   # Phone — filled by contacts.py
                    "",   # Enrich? — user marks Y
                    person.get("apollo_id", ""),
                ]
                for ci, (val, (_, __, align)) in enumerate(zip(vals, DM_COLS), 1):
                    cell = ws.cell(row, ci, val)
                    _data_style(cell, bold=(ci == 2), align=align, row_idx=sno)

    ws.freeze_panes = "B9"
    ws.auto_filter.ref = f"A8:{get_column_letter(len(DM_COLS))}8"

    if sno == 0:
        ws.cell(9, 2, "No Apollo results found — check APOLLO_API_KEY in .env").font = Font(name="Arial", size=10, italic=True, color="FF0000")

# ── Excel: assemble workbook ──────────────────────────────────────────────────
def generate_excel(strategy: Dict, enriched: Dict[str, List[Dict]], inp: Dict, path: str):
    print("\n[4/4] Generating Excel...")

    wb   = Workbook()
    wb.remove(wb.active)
    side = "sell-side" if inp["is_sellside"] else "buy-side"
    title = f"{inp['sub_sector']} — {side} mapping"

    _snapshot_sheet(wb, strategy, inp)

    for cat in strategy["categories"]:
        cos = enriched.get(cat["name"], [])
        if not cos:
            continue
        _mapping_sheet(wb, cat["name"], cos, title, inp["client"], inp["date"])

    _decision_makers_sheet(wb, enriched, inp["date"])

    wb.save(path)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not SERP_API_KEY:
        sys.exit("ERROR: SERP_API_KEY not set in .env")

    inp      = collect_inputs()
    strategy = generate_strategy(inp)

    print(f"\nFilters applied:")
    print(f"  Target geography : {inp['target_geography']}")
    print(f"  Revenue range    : {inp['revenue_range']}")
    print(f"  Company types    : {', '.join(inp['company_types'])}")
    print(f"  Must-have attrs  : {inp['specific_attrs']}")
    print(f"  Seeking          : {inp['what_looking_for']}")
    print(f"  Volume target    : {inp['count_label']} ({inp['per_cat_cap']} per category)")

    print(f"\nMapping categories:")
    for i, c in enumerate(strategy["categories"], 1):
        print(f"  {i}. {c['name']}")

    print("\n[2/4] Building company universe (Claude + search):")
    discovered = discover_companies(strategy["categories"], inp)
    enriched   = enrich_all(discovered, inp)

    # Output to Downloads
    date_tag = datetime.today().strftime("%d%b%y")
    filename = f"Accomplir - {inp['client']} - Mapping - {date_tag}.xlsx"
    out_path = os.path.join(os.path.expanduser("~/Downloads"), filename)

    generate_excel(strategy, enriched, inp, out_path)

    total = sum(len(v) for v in enriched.values())
    print(f"\n{'═'*62}")
    print(f"  Done  —  {total} companies across {len(enriched)} categories")
    print(f"  Saved: {out_path}")
    print("═"*62 + "\n")


if __name__ == "__main__":
    main()
