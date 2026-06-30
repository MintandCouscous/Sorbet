#!/usr/bin/env python3
"""
contacts.py — Step 3: Pull emails + phones for shortlisted decision makers.

Usage:
    python3 contacts.py "path/to/Accomplir - Client - Mapping - Date.xlsx"

How it works:
    1. Opens the "Decision Makers" sheet in the mapping file
    2. For every row where the "Enrich? (Y to pull email/phone)" column = Y
       (and Full Name is filled in), calls Apollo People Match
    3. Writes email + phone back into the same sheet
    4. Saves the file

Cost: 1 Apollo credit per person enriched.
"""

import os
import sys
import time
import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

APOLLO_API_KEY = os.getenv("APOLLO_API_KEY", "")

if not APOLLO_API_KEY:
    sys.exit("ERROR: APOLLO_API_KEY not set in .env")

if len(sys.argv) < 2:
    sys.exit("Usage: python3 contacts.py <path_to_mapping.xlsx>")

xlsx_path = sys.argv[1]

# ── Column positions in Decision Makers sheet (1-indexed) ────────────────────
COL_SNO         = 1
COL_COMPANY     = 2
COL_CATEGORY    = 3
COL_FIRST_NAME  = 4
COL_LAST_OBF    = 5
COL_FULL_NAME   = 6
COL_TITLE       = 7
COL_HAS_EMAIL   = 8
COL_HAS_PHONE   = 9
COL_EMAIL       = 10
COL_PHONE       = 11
COL_ENRICH      = 12
COL_APOLLO_ID   = 13

HEADER_ROW = 8
DATA_START  = 9

def apollo_match(first_name: str, full_name: str, company: str, apollo_id: str) -> dict:
    """
    Reveal email + phone via Apollo. Costs 1 credit per match.
    Uses Apollo ID directly — no manual name completion needed.
    Falls back to name + company if no ID.
    """
    # Always include name + company — match endpoint needs context even with an ID
    parts = (full_name or first_name).strip().split(" ", 1)
    payload = {
        "first_name":             parts[0],
        "last_name":              parts[1] if len(parts) > 1 else "",
        "organization_name":      company,
        "reveal_personal_emails": True,
    }
    if apollo_id:
        payload["id"] = apollo_id

    try:
        r = requests.post(
            "https://api.apollo.io/api/v1/people/match",
            headers={"x-api-key": APOLLO_API_KEY, "Content-Type": "application/json"},
            json=payload,
            timeout=15,
        )
        if not r.ok:
            print(f"    [Apollo {r.status_code}] {r.text[:300]}")
            return {"email": "", "phone": "", "full_name": ""}
        r.raise_for_status()
        person = r.json().get("person") or {}
        return {
            "email":      person.get("email") or "",
            "phone":      person.get("sanitized_phone") or "",
            "full_name":  f"{person.get('first_name','')} {person.get('last_name','')}".strip(),
        }
    except Exception as e:
        print(f"    [Apollo match error] {e}")
        return {"email": "", "phone": "", "full_name": ""}


def main():
    print(f"\nLoading: {xlsx_path}")
    wb = load_workbook(xlsx_path)

    if "Decision Makers" not in wb.sheetnames:
        sys.exit("ERROR: 'Decision Makers' sheet not found in this file.")

    ws = wb["Decision Makers"]

    to_enrich = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        enrich_flag = str(row[COL_ENRICH - 1].value or "").strip().upper()
        if enrich_flag != "Y":
            continue
        full_name  = str(row[COL_FULL_NAME  - 1].value or "").strip()
        first_name = str(row[COL_FIRST_NAME - 1].value or "").strip()
        company    = str(row[COL_COMPANY    - 1].value or "").strip()
        apollo_id  = str(row[COL_APOLLO_ID  - 1].value or "").strip()
        already_enriched = str(row[COL_EMAIL - 1].value or "").strip()
        if already_enriched:
            print(f"  Skipping (already enriched): {first_name} @ {company}")
            continue
        to_enrich.append((row, first_name, full_name, company, apollo_id))

    if not to_enrich:
        print("No rows marked for enrichment (put Y in the 'Enrich?' column).")
        return

    print(f"\nEnriching {len(to_enrich)} contacts — 1 Apollo credit each...\n")

    done = 0
    for row, first_name, full_name, company, apollo_id in to_enrich:
        done += 1
        display_name = full_name or first_name
        print(f"  [{done}/{len(to_enrich)}] {display_name} @ {company}", end=" ... ", flush=True)

        result = apollo_match(first_name, full_name, company, apollo_id)

        row[COL_EMAIL - 1].value = result["email"]
        row[COL_PHONE - 1].value = result["phone"]
        if result.get("full_name") and not row[COL_FULL_NAME - 1].value:
            row[COL_FULL_NAME - 1].value = result["full_name"]

        status = []
        if result["email"]: status.append(f"email: {result['email']}")
        if result["phone"]: status.append(f"phone: {result['phone']}")
        print(", ".join(status) if status else "no data found")

        time.sleep(0.5)   # Apollo rate limiting

    wb.save(xlsx_path)
    print(f"\n✓ Saved back to: {xlsx_path}")
    print(f"  {done} contacts processed ({sum(1 for r,*_ in to_enrich)} attempted)")


if __name__ == "__main__":
    main()
